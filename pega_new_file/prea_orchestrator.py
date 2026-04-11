#!/usr/bin/env python3
"""
prea_orchestrator.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PREA — Master Pipeline Orchestrator

Runs the full PREA pipeline from .bin files to FRD, or individual stages.

Commands:
    full      — Complete pipeline: extract → graph → trace → FRD
    extract   — Binary extraction only
    graph     — Build dependency graph from extracted rules
    trace     — Trace flow execution paths
    frd       — Generate FRD document
    decomm    — Generate decommission scorecard
    status    — Show status of all pipeline outputs

Usage:
    # Full pipeline (most common)
    python prea_orchestrator.py full \\
        --bin-dir ./exports \\
        --manifest manifest.json \\
        --output-dir ./prea_output \\
        --system-name "KYC Platform" \\
        --api-key sk-ant-...

    # Extract only
    python prea_orchestrator.py extract \\
        --bin-dir ./exports --manifest manifest.json

    # Decommission scorecard
    python prea_orchestrator.py decomm \\
        --rules ./prea_output/rules_extracted.json \\
        --graph ./prea_output/rule_graph.json \\
        --output ./prea_output/decomm_scorecard.xlsx
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import argparse
import json
import logging
import os
import subprocess
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any

log = logging.getLogger("prea.orchestrator")
logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s  %(levelname)-8s  %(message)s",
                    datefmt="%H:%M:%S")

SCRIPTS_DIR = Path(__file__).parent

# ── Stage Runner ─────────────────────────────────────────────────────────────

def run_stage(script: str, args_list: List[str], stage_name: str) -> bool:
    """Run a PREA pipeline stage via subprocess."""
    cmd = [sys.executable, str(SCRIPTS_DIR / script)] + args_list
    log.info("▶ Running stage: %s", stage_name)
    log.debug("  Command: %s", " ".join(cmd))
    t0 = time.time()
    result = subprocess.run(cmd, capture_output=False, text=True)
    elapsed = time.time() - t0
    if result.returncode == 0:
        log.info("✓ %s completed in %.1fs", stage_name, elapsed)
        return True
    else:
        log.error("✗ %s failed (exit %d)", stage_name, result.returncode)
        return False


# ── Decommission Scorer ───────────────────────────────────────────────────────

def compute_decomm_score(rule: Dict, graph_data: Optional[Dict]) -> Dict:
    """
    Compute a decommission readiness score for a rule.

    Score components (0-100, higher = safer to decommission):
      - No dependents:        +40 points
      - Low coverage:         +20 points  (coverage < 20%)
      - Deprecated status:    +20 points
      - No flow steps:        +10 points  (not a core flow)
      - Many alternatives:     +10 points (override exists in higher layer)

    Risk level:
      - Score >= 75: LOW risk
      - Score 45-74: MEDIUM risk
      - Score <  45: HIGH risk
    """
    score = 0
    factors = []

    name   = rule.get("name", "")
    rtype  = rule.get("rule_type", "")
    status = rule.get("status", "Active")
    notes  = rule.get("notes", "")
    deps   = rule.get("dependencies", [])
    steps  = rule.get("flow_steps", [])

    # Dependents from graph
    n_dependents = 0
    if graph_data:
        edges = graph_data.get("edges", [])
        rule_id = rule.get("rule_id", "")
        n_dependents = sum(1 for e in edges if e.get("target") == rule_id)

    # No dependents
    if n_dependents == 0:
        score += 40
        factors.append("No dependents (+40)")
    elif n_dependents < 5:
        score += 20
        factors.append(f"Low dependents: {n_dependents} (+20)")
    else:
        factors.append(f"HIGH dependents: {n_dependents} (0)")

    # Deprecated
    if status.lower() in ("deprecated", "inactive", "blocked"):
        score += 20
        factors.append("Deprecated status (+20)")

    # Leaf rule (no flow steps and no deps)
    if not steps and not deps:
        score += 10
        factors.append("Isolated — no steps or deps (+10)")

    # Override exists (higher-layer version available)
    if "overrides" in notes.lower() or "override" in notes.lower():
        score += 10
        factors.append("Override chain exists (+10)")

    # Rule type risk (flows and decision tables are harder to remove)
    type_risk = {
        "Flow": -15, "Activity": -10, "Decision Table": -5,
        "UI Section": 0, "Declare Expression": 5, "When Rule": 5,
        "Correspondence": 15, "Report Definition": 10,
    }
    type_adj = type_risk.get(rtype, 0)
    score += type_adj
    if type_adj != 0:
        factors.append(f"Rule type {rtype} ({type_adj:+d})")

    # Coverage adjustment (if available in notes)
    import re
    coverage_match = re.search(r"(\d+)%\s*coverage", notes, re.IGNORECASE)
    if coverage_match:
        cov = int(coverage_match.group(1))
        if cov < 20:
            score += 20
            factors.append(f"Low coverage {cov}% (+20)")
        elif cov > 60:
            score -= 10
            factors.append(f"High coverage {cov}% (-10)")

    score = max(0, min(100, score))   # clamp 0-100

    if score >= 75:
        risk = "LOW"
    elif score >= 45:
        risk = "MEDIUM"
    else:
        risk = "HIGH"

    return {
        "rule_id":      rule.get("rule_id", ""),
        "name":         name,
        "rule_type":    rtype,
        "layer":        rule.get("layer", ""),
        "ruleset":      rule.get("ruleset", ""),
        "status":       status,
        "n_dependents": n_dependents,
        "n_deps_out":   len(deps),
        "readiness_score": score,
        "risk_level":   risk,
        "factors":      factors,
        "recommended_action": {
            "LOW":    "Safe to decommission — create removal plan",
            "MEDIUM": "Review dependents before removal",
            "HIGH":   "Do not remove without full impact analysis",
        }[risk],
    }


def run_decomm(rules_path: Path, graph_path: Optional[Path],
               output_path: Path, top_n: int = 100) -> None:
    """Generate decommission scorecard."""
    with rules_path.open(encoding="utf-8") as f:
        data = json.load(f)
    rules = data.get("rules", data) if isinstance(data, dict) else data

    graph_data = None
    if graph_path and graph_path.exists():
        with graph_path.open(encoding="utf-8") as f:
            graph_data = json.load(f)

    log.info("Scoring %d rules for decommission readiness...", len(rules))
    scores = [compute_decomm_score(r, graph_data) for r in rules]

    # Sort by risk level and score
    risk_order = {"LOW": 0, "MEDIUM": 1, "HIGH": 2}
    scores.sort(key=lambda x: (risk_order[x["risk_level"]], -x["readiness_score"]))

    # Summary stats
    risk_counts = Counter(s["risk_level"] for s in scores)
    summary = {
        "total_scored":  len(scores),
        "low_risk":      risk_counts["LOW"],
        "medium_risk":   risk_counts["MEDIUM"],
        "high_risk":     risk_counts["HIGH"],
        "top_candidates": scores[:top_n],
    }

    out = {
        "generated":  datetime.now().isoformat(),
        "summary":    summary,
        "scorecard":  scores,
    }

    # Try to write XLSX
    try:
        _write_decomm_xlsx(scores, summary, output_path)
    except Exception as e:
        log.warning("XLSX write failed (%s) — writing JSON", e)
        json_out = output_path.with_suffix(".json")
        with json_out.open("w") as f:
            json.dump(out, f, indent=2)
        log.info("Scorecard written: %s", json_out)
        return

    log.info("Decommission scorecard written: %s", output_path)
    print(f"\n{'═'*55}")
    print(f"  Decommission Scorecard")
    print(f"  Total rules scored:  {len(scores):,}")
    print(f"  LOW risk (safe):     {risk_counts['LOW']:,}")
    print(f"  MEDIUM risk:         {risk_counts['MEDIUM']:,}")
    print(f"  HIGH risk (caution): {risk_counts['HIGH']:,}")
    print(f"{'═'*55}")


def _write_decomm_xlsx(scores: List[Dict], summary: Dict, output_path: Path) -> None:
    """Write decommission scorecard to Excel using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl --break-system-packages")

    wb = openpyxl.Workbook()

    # ─ Summary Sheet ─
    ws_sum = wb.active
    ws_sum.title = "Summary"

    NAVY  = "1B2A4A"
    GOLD  = "C9A84C"
    GREEN = "206020"
    AMBER = "B85C00"
    RED   = "8B0000"
    LIGHT = "F0F4FA"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    navy_fill    = PatternFill("solid", fgColor=NAVY)
    gold_fill    = PatternFill("solid", fgColor=GOLD)

    ws_sum["A1"] = "PREA — Decommission Scorecard"
    ws_sum["A1"].font = Font(name="Arial", bold=True, size=16, color=NAVY)
    ws_sum["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_sum["A2"].font = Font(name="Arial", size=10, color="888888")

    ws_sum["A4"] = "Risk Level";  ws_sum["B4"] = "Count"
    for cell in [ws_sum["A4"], ws_sum["B4"]]:
        cell.font = header_font; cell.fill = navy_fill

    risk_colors = {"LOW": ("E6F4EA", GREEN), "MEDIUM": ("FFF8E6", AMBER), "HIGH": ("FEECEC", RED)}
    for i, (risk, count) in enumerate([("LOW", summary["low_risk"]),
                                        ("MEDIUM", summary["medium_risk"]),
                                        ("HIGH", summary["high_risk"])]):
        r = 5 + i
        ws_sum[f"A{r}"] = risk
        ws_sum[f"B{r}"] = count
        color, font_color = risk_colors[risk]
        ws_sum[f"A{r}"].fill = PatternFill("solid", fgColor=color)
        ws_sum[f"A{r}"].font = Font(name="Arial", bold=True, color=font_color, size=11)
        ws_sum[f"B{r}"].font = Font(name="Arial", size=11)

    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 15

    # ─ Scorecard Sheet ─
    ws = wb.create_sheet("Scorecard")
    headers = ["Rule ID","Name","Type","Layer","Ruleset","Status",
               "Dependents","Deps Out","Score","Risk Level","Recommendation","Factors"]
    widths  = [12, 36, 20, 16, 22, 12, 12, 10, 8, 14, 42, 60]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font  = header_font
        cell.fill  = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    thin_border = Border(
        left=Side(style="thin", color="D0D8E8"),
        right=Side(style="thin", color="D0D8E8"),
        top=Side(style="thin", color="D0D8E8"),
        bottom=Side(style="thin", color="D0D8E8"),
    )

    for row_idx, score in enumerate(scores, 2):
        risk = score["risk_level"]
        fill_color, font_color = risk_colors.get(risk, ("FFFFFF", "000000"))
        row_data = [
            score["rule_id"], score["name"], score["rule_type"],
            score["layer"], score["ruleset"], score["status"],
            score["n_dependents"], score["n_deps_out"],
            score["readiness_score"], risk,
            score["recommended_action"],
            " | ".join(score["factors"]),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font   = Font(name="Arial", size=10,
                               color=font_color if col == 10 else "000000",
                               bold=(col == 10))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col in (2, 11, 12)))
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT)
            if col == 9:   # Score column
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


# ── Status Reporter ───────────────────────────────────────────────────────────

def report_status(output_dir: Path) -> None:
    """Report on the status of all PREA pipeline outputs."""
    files = {
        "Binary Extraction":      output_dir / "rules_extracted.json",
        "Layer Map":              output_dir / "layer_map.json",
        "Rule Graph JSON":        output_dir / "rule_graph.json",
        "Rule Graph HTML":        output_dir / "rule_graph.html",
        "Graph Metrics":          output_dir / "graph_metrics.json",
        "Decommission Scorecard": output_dir / "decomm_scorecard.xlsx",
        "FRD Document":           output_dir / "FRD_output.docx",
    }

    print(f"\n{'═'*60}")
    print(f"  PREA Pipeline Status  —  {output_dir}")
    print(f"{'═'*60}")
    for name, path in files.items():
        if path.exists():
            size = path.stat().st_size
            mtime = datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
            print(f"  ✓  {name:<30} {size/1024:>7.1f} KB   {mtime}")
        else:
            print(f"  ✗  {name:<30}  [not found]")

    # If rules extracted, show summary
    rules_path = files["Binary Extraction"]
    if rules_path.exists():
        with rules_path.open() as f:
            data = json.load(f)
        summary = data.get("summary", {})
        print(f"\n  Rules extracted:  {summary.get('total_rules', 0):,}")
        for layer, count in summary.get("by_layer", {}).items():
            print(f"    {layer:<22} {count:>8,}")
    print(f"{'═'*60}\n")


# ── Full Pipeline ─────────────────────────────────────────────────────────────

def run_full_pipeline(args) -> None:
    """Run the complete PREA pipeline end-to-end."""
    out = Path(args.output_dir)
    out.mkdir(parents=True, exist_ok=True)

    print(f"\n{'═'*60}")
    print(f"  PREA — Full Pipeline")
    print(f"  Output: {out.absolute()}")
    print(f"{'═'*60}\n")

    t_start = time.time()

    # Stage 1: Parse manifest
    if args.manifest:
        success = run_stage("pega_manifest_parser.py", [
            "--manifest", args.manifest,
            "--output",   str(out / "layer_map.json"),
            "--print",
        ], "Manifest Parsing")
        if not success:
            log.warning("Manifest parsing failed — continuing without layer map")

    # Stage 2: Binary extraction
    extract_args = ["--output", str(out / "rules_extracted.json")]
    if args.bin_dir:
        extract_args += ["--input-dir", args.bin_dir]
    if args.bin_file:
        extract_args += ["--input-file", args.bin_file]
    if args.manifest:
        extract_args += ["--manifest", args.manifest]
    if args.verbose:
        extract_args.append("--verbose")

    success = run_stage("pega_bin_extractor.py", extract_args, "Binary Extraction")
    if not success:
        log.error("Binary extraction failed — aborting pipeline")
        return

    rules_path = out / "rules_extracted.json"

    # Stage 3: Graph building
    run_stage("pega_graph_builder.py", [
        "--rules",          str(rules_path),
        "--output-graph",   str(out / "rule_graph.json"),
        "--output-html",    str(out / "rule_graph.html"),
        "--output-metrics", str(out / "graph_metrics.json"),
        "--app-name",       args.system_name,
    ], "Dependency Graph")

    # Stage 4: Flow tracing (top flows only in full pipeline)
    flow_trace_dir = out / "flow_traces"
    run_stage("pega_flow_tracer.py", [
        "--rules",      str(rules_path),
        "--all-flows",
        "--output-dir", str(flow_trace_dir),
        "--max-depth",  "6",
    ], "Flow Tracing")

    # Stage 5: Decommission scorecard
    run_decomm(
        rules_path  = rules_path,
        graph_path  = out / "rule_graph.json",
        output_path = out / "decomm_scorecard.xlsx",
    )

    # Stage 6: FRD generation
    frd_args = [
        "--rules",       str(rules_path),
        "--graph",       str(out / "rule_graph.json"),
        "--output",      str(out / "FRD_output.docx"),
        "--system-name", args.system_name,
        "--version",     args.version,
    ]
    api_key = args.api_key or os.environ.get("ANTHROPIC_API_KEY")
    if api_key:
        frd_args += ["--api-key", api_key]
    else:
        frd_args.append("--no-ai")

    run_stage("pega_frd_writer.py", frd_args, "FRD Generation")

    # Final status
    elapsed = time.time() - t_start
    report_status(out)
    print(f"  Pipeline completed in {elapsed:.0f}s\n")


# ── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="PREA Master Orchestrator",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    sub = parser.add_subparsers(dest="command", required=True)

    # ─ full ─
    p_full = sub.add_parser("full", help="Run complete pipeline")
    p_full.add_argument("--bin-dir",     help="Directory of .bin files")
    p_full.add_argument("--bin-file",    help="Single .bin file")
    p_full.add_argument("--manifest",    help="Manifest JSON path")
    p_full.add_argument("--output-dir",  default="./prea_output")
    p_full.add_argument("--system-name", default="Pega Application")
    p_full.add_argument("--version",     default="1.0")
    p_full.add_argument("--api-key",     help="Anthropic API key")
    p_full.add_argument("--verbose",     action="store_true")

    # ─ extract ─
    p_ext = sub.add_parser("extract", help="Binary extraction only")
    p_ext.add_argument("--bin-dir")
    p_ext.add_argument("--bin-file")
    p_ext.add_argument("--manifest")
    p_ext.add_argument("--output",   default="rules_extracted.json")
    p_ext.add_argument("--verbose",  action="store_true")

    # ─ graph ─
    p_graph = sub.add_parser("graph", help="Build dependency graph")
    p_graph.add_argument("--rules",         required=True)
    p_graph.add_argument("--output-graph",  default="rule_graph.json")
    p_graph.add_argument("--output-html",   default="rule_graph.html")
    p_graph.add_argument("--layer-filter")
    p_graph.add_argument("--app-name",      default="Pega Application")

    # ─ trace ─
    p_trace = sub.add_parser("trace", help="Trace flow execution")
    p_trace.add_argument("--rules",      required=True)
    p_trace.add_argument("--flow",       help="Specific flow name")
    p_trace.add_argument("--all-flows",  action="store_true")
    p_trace.add_argument("--output",     default="flow_trace.json")
    p_trace.add_argument("--output-dir", default="./flow_traces")

    # ─ frd ─
    p_frd = sub.add_parser("frd", help="Generate FRD document")
    p_frd.add_argument("--rules",       required=True)
    p_frd.add_argument("--graph")
    p_frd.add_argument("--output",      default="FRD_output.docx")
    p_frd.add_argument("--system-name", default="Pega Application")
    p_frd.add_argument("--version",     default="1.0")
    p_frd.add_argument("--api-key")
    p_frd.add_argument("--no-ai",       action="store_true")

    # ─ decomm ─
    p_decomm = sub.add_parser("decomm", help="Decommission scorecard")
    p_decomm.add_argument("--rules",   required=True)
    p_decomm.add_argument("--graph")
    p_decomm.add_argument("--output",  default="decomm_scorecard.xlsx")
    p_decomm.add_argument("--top-n",   type=int, default=100)

    # ─ status ─
    p_status = sub.add_parser("status", help="Show pipeline output status")
    p_status.add_argument("--output-dir", default="./prea_output")

    args = parser.parse_args()

    if args.command == "full":
        if not args.bin_dir and not args.bin_file:
            parser.error("Provide --bin-dir or --bin-file")
        run_full_pipeline(args)

    elif args.command == "extract":
        extract_args = ["--output", args.output]
        if args.bin_dir:   extract_args += ["--input-dir",  args.bin_dir]
        if args.bin_file:  extract_args += ["--input-file", args.bin_file]
        if args.manifest:  extract_args += ["--manifest",   args.manifest]
        if args.verbose:   extract_args.append("--verbose")
        run_stage("pega_bin_extractor.py", extract_args, "Binary Extraction")

    elif args.command == "graph":
        graph_args = [
            "--rules",       args.rules,
            "--output-graph",args.output_graph,
            "--output-html", args.output_html,
            "--app-name",    args.app_name,
        ]
        if args.layer_filter: graph_args += ["--layer-filter", args.layer_filter]
        run_stage("pega_graph_builder.py", graph_args, "Graph Building")

    elif args.command == "trace":
        trace_args = ["--rules", args.rules]
        if args.flow:       trace_args += ["--flow", args.flow, "--output", args.output]
        if args.all_flows:  trace_args += ["--all-flows", "--output-dir", args.output_dir]
        run_stage("pega_flow_tracer.py", trace_args, "Flow Tracing")

    elif args.command == "frd":
        frd_args = [
            "--rules",       args.rules,
            "--output",      args.output,
            "--system-name", args.system_name,
            "--version",     args.version,
        ]
        if args.graph:   frd_args += ["--graph", args.graph]
        if args.no_ai:   frd_args.append("--no-ai")
        api_key = args.api_key or os.environ.get("ANTHROPIC_API_KEY")
        if api_key: frd_args += ["--api-key", api_key]
        run_stage("pega_frd_writer.py", frd_args, "FRD Generation")

    elif args.command == "decomm":
        run_decomm(
            rules_path  = Path(args.rules),
            graph_path  = Path(args.graph) if args.graph else None,
            output_path = Path(args.output),
            top_n       = args.top_n,
        )

    elif args.command == "status":
        report_status(Path(args.output_dir))


if __name__ == "__main__":
    main()
